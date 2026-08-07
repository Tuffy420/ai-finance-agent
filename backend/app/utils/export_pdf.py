"""
Export Utilities for PDF, CSV, and Excel Reports
"""

import io
import csv
from typing import List, Dict, Any
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportExportUtils:
    @staticmethod
    def generate_pdf(user_name: str, period: str, transactions: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """
        Generate a PDF financial statement.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []
        styles = getSampleStyleSheet()

        # Title & Header
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#7B61FF'), spaceAfter=8)
        elements.append(Paragraph("FinPilot AI — Financial Statement", title_style))
        elements.append(Paragraph(f"<b>Account Holder:</b> {user_name} | <b>Statement Period:</b> {period}", styles['Normal']))
        elements.append(Spacer(1, 14))

        # Summary Table
        summary_data = [
            ["Total Income", "Total Spending", "Net Savings", "Savings Rate"],
            [
                f"+${summary.get('total_income', 12450.00):,.2f}",
                f"${summary.get('total_spending', 3450.75):,.2f}",
                f"${summary.get('net_savings', 8999.25):,.2f}",
                f"{summary.get('savings_rate_percent', 72.3)}%"
            ]
        ]
        sum_table = Table(summary_data, colWidths=[130, 130, 130, 130])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F162E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#5EA1FF')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ]))
        elements.append(sum_table)
        elements.append(Spacer(1, 18))

        # Transactions Table
        elements.append(Paragraph("<b>Recent Ledger Activity</b>", styles['Heading3']))
        tx_data = [["Date", "Merchant", "Category", "Method", "Amount"]]
        for t in transactions[:20]:
            tx_data.append([
                str(t.get("date", "Today")),
                str(t.get("merchant", "")),
                str(t.get("category", "")),
                str(t.get("payment_method", "")),
                f"${abs(float(t.get('amount', 0))):.2f}"
            ])
            
        tx_table = Table(tx_data, colWidths=[80, 160, 100, 90, 90])
        tx_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B61FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ]))
        elements.append(tx_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_csv(transactions: List[Dict[str, Any]]) -> str:
        """
        Generate RFC-compliant CSV text.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Merchant", "Category", "Amount", "Currency", "PaymentMethod", "Date", "Source"])
        for t in transactions:
            writer.writerow([
                t.get("id", ""),
                t.get("merchant", ""),
                t.get("category", ""),
                t.get("amount", 0.0),
                t.get("currency", "USD"),
                t.get("payment_method", "UPI"),
                str(t.get("transaction_date", "")),
                t.get("source", "sms")
            ])
        return output.getvalue()

    @staticmethod
    def generate_excel(transactions: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """
        Generate styled multi-tab Excel spreadsheet using openpyxl.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger Activity"

        # Headers
        headers = ["Transaction ID", "Merchant", "Category", "Amount", "Currency", "Payment Method", "Date", "Source"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="7B61FF", end_color="7B61FF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for t in transactions:
            ws.append([
                str(t.get("id", "")),
                str(t.get("merchant", "")),
                str(t.get("category", "")),
                float(t.get("amount", 0.0)),
                str(t.get("currency", "USD")),
                str(t.get("payment_method", "UPI")),
                str(t.get("transaction_date", "")),
                str(t.get("source", "sms"))
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
